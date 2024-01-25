# deepsort
from main.utils_ds.parser import get_config
from main.utils_ds.draw import draw_boxes
from main.deep_sort import build_tracker
import datetime
import cv2
import argparse
import os
import time
import numpy as np
import warnings
import torch
import pickle
import torch.backends.cudnn as cudnn
import mysql.connector
from PIL import Image
import shutil
import sys
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtChart import *
from PyQt5.QtCore import *
from PyQt5.QtPrintSupport import *
from threading import Thread
from UI import Ui_MainWindow  # 導入Ui_MainWindow (UI檔案)
from pop import Ui_Dialog
from facenet_pytorch import InceptionResnetV1
# 匯出
import csv
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import *
from PyPDF2 import *
from reportlab.lib.units import inch

# 人臉偵測器
from facenet_pytorch import MTCNN

import sys
import gc
import tracemalloc
import threading


# 寄email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

currentUrl = os.path.dirname(__file__)
sys.path.append(os.path.abspath(os.path.join(currentUrl, 'yolov5')))
cudnn.benchmark = True

#######################################  deepsort  ####################################
ui = Ui_MainWindow()
#######################################  設定參數  ################
update_count_in=5   # 連續偵測的偵數
update_count_out=5  # 連續偵測的偵數
MTCNN_num=3     #MTCNN參數
face_num=0.78   #人臉相似度的閾值(利用歐氏距離去計算)
#######################################
# *********資料庫初始化***********
# 建立MySQL連線
conn = mysql.connector.connect(
    host='127.0.0.1',
    user='root',
    password='daisy218',
    database='data'  
)
# 建立游標
cursor = conn.cursor()

table_name = "today"
# 刪除資料表（如果存在的話）
try:
    cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
    print(f"{table_name} delete successfully")
except mysql.connector.Error as err:
    print(f"error：{err}")

# 建立新的資料表
create_table_query = """
CREATE TABLE IF NOT EXISTS today (
    id INT AUTO_INCREMENT PRIMARY KEY,
    feature BLOB,
    name VARCHAR(255) DEFAULT NULL,  
    number VARCHAR(255) DEFAULT NULL,
    image_path VARCHAR(255) NOT NULL,
    in_time DATETIME    
)
 """
cursor.execute(create_table_query)

#存通知對象的e-mail、名字資料表
create_table_query = """
CREATE TABLE IF NOT EXISTS notify (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255) DEFAULT NULL,  
    email VARCHAR(255) NOT NULL
)
"""
cursor.execute(create_table_query)

create_table_query = """
CREATE TABLE IF NOT EXISTS history (
    id INT AUTO_INCREMENT PRIMARY KEY,
    feature BLOB,
    name VARCHAR(255) DEFAULT NULL,
    number VARCHAR(255) DEFAULT NULL,
    image_path VARCHAR(255)DEFAULT NULL,
    in_time DATETIME,
    out_time DATETIME
)
"""
cursor.execute(create_table_query)

create_table_query = """
CREATE TABLE IF NOT EXISTS people_count (
    date DATETIME,
    count INT
)
"""
cursor.execute(create_table_query)

# 特殊名單
special_encoding = []   # 特徵向量
special_name = []       # 名稱
special_number = []     # 學號

# 執行 SQL 查詢
cursor.execute("SELECT * FROM special_list")
rows = cursor.fetchall()

for row in rows:
    # 提取名稱和特徵向量
    name = row[1]
    number = row[2]
    feature1 = pickle.loads(row[3]) # 將BLOB轉成特徵向量numpy(list)
    feature2 = pickle.loads(row[5])
    feature3 = pickle.loads(row[7])
    feature4 = pickle.loads(row[9])
    feature5 = pickle.loads(row[11])
    feature6 = pickle.loads(row[13])

# 添加特徵向量和名稱到列表
    special_encoding.extend(
        [feature1, feature2, feature3, feature4, feature5, feature6])
    special_name.extend([name, name, name, name, name, name])
    special_number.extend([number, number, number, number, number, number])

conn.commit()

# 通知名單
notify_email = []
notify_name = []

cursor.execute("SELECT * FROM notify")
rows = cursor.fetchall()

for row in rows:
    name = row[1]
    email = row[2]

    notify_email.append(email)
    notify_name.append(name)

conn.commit()


# 建立當日圖庫
current_time = datetime.datetime.now().strftime('%Y-%m-%d')
today_file = os.path.join('./main/crop_img', current_time)
if not os.path.exists(today_file):
    os.makedirs(today_file)
######################################################################################


class VideoTracker(object):
    def __init__(self, args):
        print('Initialize DeepSORT & YOLO-V5')
        super().__init__()
        tracemalloc.start()
        self.timer_clear()
        # ******************************* 初始化 ****************************************
        self.args = args
        self.scale = args.scale                         # 2
        self.margin_ratio = args.margin_ratio           # 0.2
        self.frame_interval = args.frame_interval       # 頻率
        self.stopped = False  # 暫停

        self.device = torch.device(
            'cuda:0' if torch.cuda.is_available() else 'cpu')
        self.half = self.device.type != 'cpu'  # 只有CUDA支持半精度
        self.history_in = {}  # 用於存儲每個人的進入和離開事件
        self.history_out = {}

        self.total_people_incount = 0  # 歷史人數
        self.total_people_outcount = 0  # 歷史人數

        self.facenet = InceptionResnetV1(pretrained='vggface2').eval()
        ##
        #sender_email = 'kitty46978@gmail.com'  # 你的email
        #password = 'mrxm ftjv sqyg llvm'  # 應用程式密碼
        self.sender_email = '09360085@me.mcu.edu.tw'  # 你的email
        password = 'cvur zzrm pvsk tgzh'  # 應用程式密碼
        # 建立 MIMEMultipart 和 smtp 連接
        self.msg = MIMEMultipart()
        self.smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
        self.smtp_server.starttls()
        self.smtp_server.login(self.sender_email , password)

        # 捕捉鏡頭 ****************

        if args.cam != -1:
            print("Using webcam " + str(args.cam))
            self.entry_vdo = cv2.VideoCapture(args.cam)
            self.entry_vdo.set(6, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'))
        if args.cam1 != -1:
            print("Using webcam " + str(args.cam1))
            self.exit_vdo = cv2.VideoCapture(args.cam1)
            self.exit_vdo.set(6, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'))
        else:
            self.vdo = cv2.VideoCapture()

        # ***************************** 初始化 DeepSORT *********************************
        cfg = get_config()
        cfg.merge_from_file(args.config_deepsort)

        use_cuda = self.device.type != 'cpu' and torch.cuda.is_available()
        self.deepsort_in = build_tracker(cfg, use_cuda=use_cuda)#進館
        self.deepsort_out = build_tracker(cfg, use_cuda=use_cuda)#離館
        # ***************************** 初始化人臉偵測 **********************************
        self.face_detector = MTCNN(keep_all=True, device=self.device)

        print('Done..')

        if self.device == 'cpu':
            warnings.warn(
                "Running in cpu mode which maybe very slow!", UserWarning)

        # *******************************************************************************
    def __enter__(self):
        # ************************* 從攝像頭加載視頻 *************************
        if self.args.cam != -1:
            print('攝像頭模式...')
            ret, self.entry_frame = self.entry_vdo.read()
            assert ret, "Error: Camera error"
            self.im_width = int(self.entry_vdo.get(cv2.CAP_PROP_FRAME_WIDTH))
            self.im_height = int(self.entry_vdo.get(cv2.CAP_PROP_FRAME_HEIGHT))
        if self.args.cam1 != -1:
            print('攝像頭模式...')
            ret, self.exit_frame = self.exit_vdo.read()
            assert ret, "Error: Camera error"
            self.im_width = int(self.exit_vdo.get(cv2.CAP_PROP_FRAME_WIDTH))
            self.im_height = int(self.exit_vdo.get(cv2.CAP_PROP_FRAME_HEIGHT))

        # ************************* 從文件加載視頻 *************************

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.smtp_server.quit()
        self.garbage_collection()
        self.timer.cancel()
        tracemalloc.stop()
        if self.entry_vdo is not None:
            self.entry_vdo.release()
        if self.exit_vdo is not None:
            self.exit_vdo.release()
        if exc_type:
            print(exc_type, exc_value, exc_traceback)
            
    def garbage_collection(self):
        gc.collect()
        new_stats = tracemalloc.get_traced_memory()
        print(f"Memory usage after garbage collection: {new_stats[0] / 10**6} MB")
        
    def timer_clear(self):
        self.garbage_collection()
        # 每3分鐘執行一次
        self.timer = threading.Timer(180, self.timer_clear)
        self.timer.start()

    def run(self):
        entry_yolo_time, entry_sort_time, entry_avg_fps = [], [], []
        exit_yolo_time, exit_sort_time, exit_avg_fps = [], [], []
        entry_idx_frame, exit_idx_frame = 0, 0
        entry_last_out, exit_last_out = None, None
        while not self.stopped and ((self.entry_vdo is not None) or (self.exit_vdo is not None)):
            _, self.entry_img0 = self.entry_vdo.read()
            _, self.exit_img0 = self.exit_vdo.read()
            if self.entry_img0 is not None:
                entry_t0 = time.time()
                if entry_idx_frame % self.frame_interval == 0:
                    entry_outputs, entry_yt, entry_st = self.image_track_entry(self.entry_img0)
                    entry_last_out = entry_outputs
                    entry_yolo_time.append(entry_yt)
                    entry_sort_time.append(entry_st)
                    print('Entry Frame %d Done. Det-time:(%.3fs) SORT-time:(%.3fs)' %
                          (entry_idx_frame, entry_yt, entry_st))
                else:
                    entry_outputs = entry_last_out
                entry_t1 = time.time()
                print(entry_t1 - entry_t0)
                entry_avg_fps.append(entry_t0 - entry_t1)

                if len(entry_outputs) > 0:  #entry_outputs有x1,y1,x2,y2,track_id
                    entry_bbox_xyxy = entry_outputs[:, :4]  #x1,y1,x2,y2
                    entry_identities = entry_outputs[:, -1]  #track_id

                    people = self.update_history_entry(entry_outputs, self.entry_img0)  # 更新人數
                    self.entry_img0 = draw_boxes(
                        self.entry_img0, entry_bbox_xyxy)  # BGR
                    for out, id in zip(entry_bbox_xyxy, entry_identities):
                        pp = False  # 是否為特定名單
                        feature = self.encoding(self.entry_img0, out)
                        entry_num = self.recognize_face(feature, special_encoding, special_number)
                        if entry_num in special_number:
                            entry_name = special_name[special_number.index(entry_num)]
                            cv2.putText(self.entry_img0, entry_name, (
                                out[0] + 6, out[3] - 6), cv2.FONT_HERSHEY_DUPLEX, 2, (255, 255, 255), 2)
                            pp = True
                        else:
                            entry_name = None
                                
    
                    
                    if people:
                        # 存入資料庫
                        feature_np = feature.detach().numpy()
                        # 獲取當前時間
                        current_time_in = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                        cursor.execute("UPDATE today SET feature = %s,in_time = %s,name=%s,number=%s  WHERE id = %s", (pickle.dumps(
                            feature_np), current_time_in, entry_name, entry_num, int(id)))
                        conn.commit()

                        if pp and entry_name is not None:
                            self.send_email(entry_name)
    

            if self.exit_img0 is not None:
                exit_t0 = time.time()
                if exit_idx_frame % self.args.frame_interval == 0:
                    exit_outputs, exit_yt, exit_st=None,None,None
                    exit_outputs, exit_yt, exit_st = self.image_track_exit(self.exit_img0)
                    exit_last_out = exit_outputs
                    exit_yolo_time.append(exit_yt)
                    exit_sort_time.append(exit_st)
                    print('Exit Frame %d Done. Det-time:(%.3fs) SORT-time:(%.3fs)' %
                          (exit_idx_frame, exit_yt, exit_st))
                else:
                    exit_outputs = exit_last_out
                exit_t1 = time.time()
                print(exit_t1 - exit_t0)
                exit_avg_fps.append(exit_t0 - exit_t1)
                
                if len(exit_outputs)==0:
                    self.video_player(self.entry_img0, self.exit_img0)
                    continue

                exit_bbox_xyxy = exit_outputs[:, :4]

                self.update_history_exit(exit_outputs)  # 更新人數
                self.exit_img0 = draw_boxes(self.exit_img0, exit_bbox_xyxy)  # BGR
                

                for exit_out in exit_bbox_xyxy:
                    exit_feature = self.encoding(self.exit_img0, exit_out)
                    exit_num = self.recognize_face(
                        exit_feature, special_encoding, special_number)
                    if exit_num in special_number:
                        exit_name = special_name[special_number.index(exit_num)]
                        cv2.putText(self.exit_img0, exit_name, (
                        exit_out[0] + 6, exit_out[3] - 6), cv2.FONT_HERSHEY_DUPLEX, 2, (255, 255, 255), 2)

                    exist = self.recognize_face_db(exit_feature)

                    if exist:
                        query = "SELECT * FROM today WHERE id = %s"  # 查詢
                        cursor.execute(query, (int(exist),))
                        data_to_save = cursor.fetchone()
                        if data_to_save:
                            current_time_out = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                            insert_query = "INSERT INTO history (feature,name,number,image_path,in_time,out_time) VALUES (%s,%s, %s, %s, %s, %s)"
                            cursor.execute(insert_query, (data_to_save[1], data_to_save[2], data_to_save[3], data_to_save[4], data_to_save[5], current_time_out))
                            conn.commit()
                            delete_query = "DELETE FROM today WHERE id=%s"  # 刪除
                            cursor.execute(
                                delete_query, (data_to_save[0],))
                            conn.commit()

                            cv2.putText(self.exit_img0, "Exit Person", (
                                exit_out[0] + 6, exit_out[3] - 6), cv2.FONT_HERSHEY_DUPLEX, 5, (30,144,255), 10)

                            
                            

            if self.entry_vdo is not None:
                if self.args.display:
                    self.video_player(self.entry_img0, self.exit_img0)
                    entry_idx_frame += 1
                    exit_idx_frame += 1

        if self.entry_vdo is not None:
            print('Entry Avg Det time (%.3fs), SORT time (%.3fs) per frame' % (sum(entry_yolo_time) / len(entry_yolo_time),
                                                                               sum(entry_sort_time) / len(entry_sort_time)))
        if self.exit_vdo is not None:
            print('Exit Avg Det time (%.3fs), SORT time (%.3fs) per frame' % (sum(exit_yolo_time) / len(exit_yolo_time),
                                                                              sum(exit_sort_time) / len(exit_sort_time)))
        

    def stop(self):
        self.stopped = True
        self.entry_vdo.release() 
        self.exit_vdo.release()
        self.smtp_server.quit()
        self.timer.cancel()
        gc.collect()
    
    def encoding(self,img, face):
        x1, y1, x2, y2 = face
        face_image = img[y1:y2, x1:x2]
        # 檢查 face_image 是否為空
        if face_image is None:
            return None
        # 檢查 face_image 是否為非空矩陣
        if face_image.size == 0:
            return None
        # 將人臉圖像轉換為FaceNet所需的格式
        face_image = cv2.cvtColor(face_image, cv2.COLOR_BGR2RGB)
        face_image = cv2.resize(face_image, (160, 160))
        face_image = (face_image - 127.5) / 128.0  # 正規化(0~1)
        # 提取人臉特徵
        face_tensor = torch.from_numpy(
            face_image.transpose(2, 0, 1)).unsqueeze(0).float()
        with torch.no_grad():
            features = self.facenet(face_tensor)
        return features

    def recognize_face(self, features, encoding, num):
        # 計算特徵向量之間的歐氏距離 
        distances = [np.linalg.norm(features.detach().numpy() - embedding) for embedding in encoding]
        # 選擇相似度最高的已知人臉
        if not distances:
            return None
        min_distance = min(distances)

        # 如果相似度高於閾值，則識別為已知人臉
        if min_distance < face_num:
            min_distance_index = distances.index(min_distance)
            recognized_name = num[min_distance_index]
            num = None
            distances=None
            return recognized_name
        return None
    def recognize_face_db(self, features):
        if features is not None:
            # 計算已知人臉特徵向量與捕捉人臉特徵向量之間的歐氏距離
            distances = []
            num=[]
            cursor.execute("SELECT id, feature FROM today")
            for row in cursor.fetchall():
                known_embedding = pickle.loads(row[1])
                # 計算歐氏距離並加入距離列表
                distance = np.linalg.norm(features.detach().numpy() - known_embedding)
                distances.append(distance)
                num.append(row[0])

            # 選擇相似度最高的已知人臉
            if not distances:
                return None
            min_distance = min(distances)

            # 如果相似度高於閾值，則識別為已知人臉
            if min_distance < face_num:
                min_distance_index = distances.index(min_distance)
                recognized_name = num[min_distance_index]
                num = None
                distances=None
                return recognized_name
        return None
    def image_track_entry(self,im0):
        # 預處理 ************************************************************
        h, w, _ = im0.shape
        img = cv2.resize(im0, (w //self.scale, h //
                        self.scale))         # 下採樣以加快速度
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # 偵測時間 *********************************************************
        t1 = time.time()
        with torch.no_grad():
            boxes, confs = self.face_detector.detect(img)   #MTCNN偵測人臉(box(x1,y1,x2,y2), conf(信心水準))

        t2 = time.time()

        # 獲取所有目標 ************************************************************

        if boxes is not None and len(boxes):
            boxes = boxes * self.scale      # x1,y1,x2,y2  回到原始圖像

            bbox_xywh = xyxy2xywh(boxes)    # (#obj, 4)     xc,yc,w,h

            # 在這裡添加邊界。只需要修訂寬度和高度
            bbox_xywh[:, 2:] = bbox_xywh[:, 2:] * (1 + self.margin_ratio)

            # ****************************** deepsort ****************************
            outputs = self.deepsort_in.update(bbox_xywh, confs, im0)    #im0: 原始圖像
            # (#ID, 5) x1,y1,x2,y2,track_ID
        else:
            outputs = torch.zeros((0, 5))

        t3 = time.time()
        return outputs, t2-t1, t3-t2

    def image_track_exit(self,im0):
        # 預處理 ************************************************************
        h, w, _ = im0.shape
        img = cv2.resize(im0, (w // self.scale, h //
                         self.scale))         # 下採樣以加快速度
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # 偵測時間 *********************************************************
        # 推論
        t1 = time.time()
        with torch.no_grad():
            boxes, confs = self.face_detector.detect(img)
            # boxes: (#obj, 4) x1,y1,x2,y2      在圖像尺寸下！
            # confs: ()

        t2 = time.time()

        # 獲取所有目標 ************************************************************

        if boxes is not None and len(boxes):
            boxes = boxes *self.scale     # x1,y1,x2,y2  回到原始圖像

            bbox_xywh = xyxy2xywh(boxes)    # (#obj, 4)     xc,yc,w,h

            # 在這裡添加邊界。只需要修訂寬度和高度
            bbox_xywh[:, 2:] = bbox_xywh[:, 2:] * (1 + self.margin_ratio)

            # ****************************** deepsort ****************************
            outputs = self.deepsort_out.update(bbox_xywh, confs, im0)
            # (#ID, 5) x1,y1,x2,y2,track_ID
        else:
            outputs = torch.zeros((0, 5))

        t3 = time.time()
        return outputs, t2-t1, t3-t2

    def update_history_entry(self, outputs, img):  # 計算人數
        bb = False
        for output in outputs:
            x1, y1, x2, y2, self.track_id = output
            # 檢查是否有新目標出現
            if self.track_id not in self.history_in:
                self.history_in[self.track_id] = 1
            else:
                self.history_in[self.track_id] += 1
                if self.history_in[self.track_id] == update_count_in:
                    self.total_people_incount += 1
                    self.crop_img(output, img)
                    bb = True   # 有人進來
        ui.entry_num.setText(str(self.total_people_incount))
        ui.inside_num.setText(str(self.total_people_incount-self.total_people_outcount))
        return bb
    def update_history_exit(self, outputs):  # 計算人數
        for output in outputs:
            x1, y1, x2, y2, self.track_id = output
            # 檢查是否有新目標出現
            if self.track_id not in self.history_out:
                self.history_out[self.track_id] = 1  # 有的話記錄並加人數
            else:
                self.history_out[self.track_id] += 1
                if self.history_out[self.track_id] == update_count_out:
                    self.total_people_outcount += 1

        ui.exit_num.setText(str(self.total_people_outcount))
        ui.inside_num.setText(str(self.total_people_incount-self.total_people_outcount))


    def crop_img(self, output, frame):
        x1, y1, x2, y2, self.track_id = output
        crop_img = frame[y1:y2, x1:x2]
        file_name = today_file+"/"+str(self.track_id)+".jpg"
        # cv2.imshow("cropped", crop_img)
        cv2.imwrite(file_name, crop_img)

        insert_query = "INSERT INTO today (id, image_path) VALUES (%s, %s)"
        cursor.execute(insert_query, (str(self.track_id), file_name))

    def video_player(self, img, img1):
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)

        desired_label_height = ui.entry_video_label.height()

        q_img = QPixmap.fromImage(
            QImage(img.data,img.shape[1], img.shape[0], img.shape[1] * 3, QImage.Format_RGB888))
        q_img1 = QPixmap.fromImage(
            QImage(img1.data,img1.shape[1], img1.shape[0], img1.shape[1] * 3, QImage.Format_RGB888))

        entry_pixmap = q_img1.scaledToHeight(
            desired_label_height, Qt.SmoothTransformation)
        exit_pixmap = q_img.scaledToHeight(
            desired_label_height, Qt.SmoothTransformation)

        ui.entry_video_label.setPixmap(entry_pixmap)
        ui.exit_video_label.setPixmap(exit_pixmap)
        QCoreApplication.processEvents()

    def send_email(self, entry_name):
        query = "SELECT number, image_path, in_time FROM today WHERE name = %s"
        cursor.execute(query, (entry_name,))
        result = cursor.fetchone()
        if not result:
            return None 
        number, image_path, in_time = result

        with open(image_path, "rb") as file:
            filecontent = file.read()

        for name, email in zip(notify_name, notify_email):
            msg = MIMEMultipart()
            msg['From'] = self.sender_email 
            msg['To'] = email
            msg['Subject'] = '特殊名單進館通知'

            body = f"""
            特殊名單進館通知
            ---------------------------

            姓名: {entry_name}
            學號/員編: {number}
            進館時間: {in_time}
            """

            msg.attach(MIMEText(body, 'plain', 'utf-8'))

            mime = MIMEImage(filecontent)
            mime.add_header('Content-Disposition',
                                'attachment', filename=f"{entry_name}.jpg")
            msg.attach(mime)

            try:
                self.smtp_server.send_message(msg)
                print(f"特殊名單進館通知成功發送至{email}")
            except Exception as e:
                print(f"特殊名單進館通知發送失敗至{email}，錯誤訊息: {e}")
        

def xyxy2xywh(x):
    # 將nx4的框從[x1, y1, x2, y2]轉換為[x, y, w, h]，其中xy1為左上角，xy2為右下角
    y = torch.zeros_like(x) if isinstance(
        x, torch.Tensor) else np.zeros_like(x)
    y[:, 0] = (x[:, 0] + x[:, 2]) / 2  # x中心
    y[:, 1] = (x[:, 1] + x[:, 3]) / 2  # y中心
    y[:, 2] = x[:, 2] - x[:, 0]  # 寬度
    y[:, 3] = x[:, 3] - x[:, 1]  # 高度
    return y


def setting():
    parser = argparse.ArgumentParser()
    # 輸入和輸出
    parser.add_argument("--frame_interval", type=int, default=1)    #條偵測頻率
    parser.add_argument('--device', default='', help='CUDA設備，例如0或0,1,2,3或cpu')
    # 只有攝像頭
    parser.add_argument("--display", action="store_true", default=True)
    parser.add_argument("--display_width", type=int, default=900)
    parser.add_argument("--display_height", type=int, default=650)
    parser.add_argument("--cam", action="store",
                        dest="cam", type=int, default="0")  # 攝像頭插槽(0 1 2)
    parser.add_argument("--cam1", action="store",
                        dest="cam1", type=int, default="1") # 攝像頭插槽(0 1 2)
    # 人臉檢測參數
    parser.add_argument("--scale", type=int, default=MTCNN_num)
    parser.add_argument("--margin_ratio", type=int, default=0.1)   

    # deepsort參數
    parser.add_argument("--config_deepsort", type=str,
                        default="./main/configs/deep_sort.yaml")

    args = parser.parse_args()
    return args


###############################  介面  ######################################


class MainWindow_controller(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        ui.setupUi(self)
        self.args = setting()  # 設定參數 (video_tracker)
        self.vt = VideoTracker(self.args)

        self.face_detector = MTCNN()
        self.setup_control()
        self.calendar_and_time()

        self.data_export()
        self.show_notify_list()
        self.show_special_list()

        self.pho = None  # 在初始化 pho
        self.video_thread = Thread(target=self.vt.run)  # video_thread 在初始化

        # 查詢輸出初始化
        self.record = None
        self.pho = None
        # 統計圖初始化
        self.chart = QChart()  # 創建 chart 物件
        self.new_layout = QVBoxLayout()
        ui.scrollAreaWidgetContents_2.setLayout(self.new_layout)
        self.chart_view = QChartView()

    def setup_control(self):
        # TODO
        # 當前時間顯示
        self.timer = QtCore.QTimer(self)      # 建立計時器
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)

        # 開啟系統按鈕
        ui.open.clicked.connect(self.start_video)
        # 查詢
        ui.search_time_btn.clicked.connect(self.search_date)
        ui.search_key_btn.clicked.connect(self.search_key)
        ui.search_img_upload_btn.clicked.connect(self.upload_img)
        ui.search_img_btn.clicked.connect(self.search_pic)
        ui.search_insert_special_btn.clicked.connect(
            self.insert_check_specialdata)
        ui.close.clicked.connect(self.clear_database)
        ui.tableWidget.itemSelectionChanged.connect(self.check)
        ui.search_only_special_ck.clicked.connect(self.specialdata)
        # 統計
        ui.draw_btn.clicked.connect(self.raddio_chart)
        ui.barchart_rbtn.clicked.connect(self.handle_date_time_change)
        ui.linechart_rbtn.clicked.connect(self.handle_date_time_change)
        ui.export_btn.clicked.connect(self.data_export_btn)
        # 管理
        ui.special_img_up_btn.clicked.connect(self.special_img)
        ui.special_imgl_up_btn.clicked.connect(self.special_img_left)
        ui.special_imgr_up_btn.clicked.connect(self.special_img_right)
        ui.special_imgm_up_btn.clicked.connect(self.special_img_mask)
        ui.special_imgml_up_btn.clicked.connect(self.special_img_mask_left)
        ui.special_imgmr_up_btn.clicked.connect(self.special_img_mask_right)
        ui.notify_insert_btn.clicked.connect(self.insert_notify_list)
        ui.notify_modify_btn.clicked.connect(self.modify_notify_list)
        ui.notify_del_btn.clicked.connect(self.delete_notify_list)
        ui.special_insert_btn.clicked.connect(self.insert_special_list)
        ui.special_modify_btn.clicked.connect(self.showDialog)
        ui.special_del_btn.clicked.connect(self.delete_special_list)
        ui.tableWidget_3.itemSelectionChanged.connect(self.check_special)

    # 刪除清館資料
    def delete_data(self, image_path):
        # 使用 image_path 來辨識要刪除哪筆資料
        try:
            delete_query = "DELETE FROM today WHERE image_path = %s"
            cursor.execute(delete_query, (image_path,))
            conn.commit()
            self.vt.total_people_outcount += 1
            enter = self.vt.total_people_incount
            out = self.vt.total_people_outcount

            ui.entry_num.setText(str(enter))
            ui.exit_num.setText(str(out))
            ui.inside_num.setText(str(enter-out))

            print(f"資料 {image_path} 已從資料庫中刪除")
        except Exception as e:
            print(f"刪除資料時發生錯誤：{e}")

    # 清館
    def clear_database(self, pho):
        query = "SELECT COUNT(*) FROM today"
        cursor.execute(query)
        result = cursor.fetchone()
        self.stop_video()

        if result[0] == 0:
            QMessageBox.information(self, '清館', '目前館內已無人，即將關閉系統!!')
            tracemalloc.stop()
            sys.exit()

        else:
            count = result[0]
            query_data = "SELECT id,name,number,image_path,in_time FROM today"
            cursor.execute(query_data)
            data = cursor.fetchall()

            data_str = '\n'.join(
                [f'ID: {row[0]}, 名字: {row[1]} ,學號: {row[2]}, 圖片路徑: {row[3]}, 進館時間: {row[4]}' for row in data])
            QMessageBox.information(self, '清館', f'館內人數還有：{count}人\n')

            # 顯示所有圖片
            for row in data:
                image_path = row[3]
                pho = cv2.imread(image_path)
                resized_image = cv2.resize(pho, (180, 180))

                # 將 BGR 格式轉換為 RGB 格式
                rgb_image = cv2.cvtColor(resized_image, cv2.COLOR_BGR2RGB)

                # 將 numpy 陣列轉換為 QImage
                q_img = QtGui.QImage(
                    rgb_image.data, rgb_image.shape[1], rgb_image.shape[0], QtGui.QImage.Format_RGB888)

                # 創建一個信息框
                msg_box = QtWidgets.QMessageBox()
                # 設定視窗標題
                msg_box.setWindowTitle("顯示圖片")
                msg_box.setIcon(QtWidgets.QMessageBox.Information)
                if row[1] is not None:
                    msg_box.setText(
                        f"這裡是 {row[2]}  {row[1]}的圖片\n進館時間: {row[4]}\n")
                else:
                    msg_box.setText(f"進館時間: {row[4]}\n")

                # 在 detailed text 中插入 QLabel 以顯示圖片
                label = QtWidgets.QLabel()
                pixmap = QtGui.QPixmap.fromImage(q_img)
                label.setPixmap(pixmap)
                label.setAlignment(QtCore.Qt.AlignCenter)

                # 創建一個“確定”按鈕
                ok_button = msg_box.addButton(QtWidgets.QMessageBox.Ok)
                ok_button.setText("確定")

                # 新增一個自定義按鈕 "刪除資料"
                delete_button = msg_box.addButton(
                    "刪除資料", QtWidgets.QMessageBox.ActionRole)
                delete_button.clicked.connect(
                    lambda _, image_path=row[3]: self.delete_data(image_path))

                # 設置 "確定" 按鈕的位置
                layout = msg_box.layout()
                layout.setContentsMargins(10, 10, 10, 10)  # 設定邊緣空白，以便更好地控制位置
                layout.addWidget(label, 0, 0, 1, 1,
                                 QtCore.Qt.AlignCenter)  # 這裡修正
                layout.addWidget(
                    ok_button, 1, 2, 1, 1, QtCore.Qt.AlignBottom | QtCore.Qt.AlignRight)  # 這裡修正

                result = msg_box.exec_()

                if result == QtWidgets.QMessageBox.Ok:
                    print("User clicked Ok")
                elif result == QtWidgets.QMessageBox.Cancel:
                    print("User clicked Cancel")

    def calendar_and_time(self):
        # 查詢
        # 創建並配置日期編輯小工具

        # 將日期設置為當前日期
        ui.search_start_dateEdit.setDate(QtCore.QDate.currentDate())
        ui.search_end_dateEdit.setDate(QtCore.QDate.currentDate().addDays(1))

        # 設置最小和最大日期
        ui.search_start_dateEdit.setMinimumDate(
            QDate.currentDate().addDays((-365)*5))
        ui.search_start_dateEdit.setMaximumDate(QDate.currentDate().addDays(1))
        ui.search_start_dateEdit.setCalendarPopup(True)

        ui.search_end_dateEdit.setMinimumDate(
            QDate.currentDate().addDays((-365)*5))
        ui.search_end_dateEdit.setMaximumDate(QDate.currentDate().addDays(1))
        ui.search_end_dateEdit.setCalendarPopup(True)

        # 開始~結束(小時/分)
        # 添加小時選項到下拉式選單
        hours = [QtCore.QTime(hour, 0)
                 for hour in range(24)]  # 使用 PyQt 的 QTime
        hour_strings = [h.toString('hh') for h in hours]  # 格式化為字符串
        ui.comboBox_start_hour.addItems(hour_strings)
        ui.comboBox_end_hour.addItems(hour_strings)

        # 添加分鐘選項到下拉式選單
        mins = [QtCore.QTime(0, minute)
                for minute in range(60)]  # 使用 PyQt 的 QTime
        min_strings = [m.toString('mm') for m in mins]  # 格式化為字符串
        ui.comboBox_start_min.addItems(min_strings)
        ui.comboBox_end_min.addItems(min_strings)

        # 統計
        # 創建並配置日期編輯小工具

        # 將日期設置為當前日期
        ui.data_start_dateEdit.setDate(QtCore.QDate.currentDate())
        ui.data_end_dateEdit.setDate(QtCore.QDate.currentDate().addDays(1))

        # 設置最小和最大日期
        ui.data_start_dateEdit.setMinimumDate(
            QDate.currentDate().addDays((-365)*5))
        ui.data_start_dateEdit.setMaximumDate(QDate.currentDate().addDays(1))
        ui.data_start_dateEdit.setCalendarPopup(True)
        self.dateChoice_data = ui.data_start_dateEdit.date()

        ui.data_end_dateEdit.setMinimumDate(
            QDate.currentDate().addDays((-365)*5))
        ui.data_end_dateEdit.setMaximumDate(QDate.currentDate().addDays(1))
        ui.data_end_dateEdit.setCalendarPopup(True)
        self.dateChoice1_data = ui.data_end_dateEdit.date()

    def show_notify_list(self):
        ui.tableWidget_2.setRowCount(0)
        row = 0
        ui.tableWidget_2.setColumnCount(4)
        ui.tableWidget_2.setHorizontalHeaderLabels(["勾選", "編號", "姓名", "email"])
        ui.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        cursor.execute("SELECT * FROM notify")
        record = cursor.fetchall()

        for id, name, email in record:
            ui.tableWidget_2.insertRow(row)

            checkbox = QCheckBox()
            ui.tableWidget_2.setCellWidget(row, 0, checkbox)

            ui.tableWidget_2.setItem(row, 1, QTableWidgetItem(str(id)))
            ui.tableWidget_2.setItem(row, 2, QTableWidgetItem(name))
            ui.tableWidget_2.setItem(row, 3, QTableWidgetItem(email))

            row += 1

    def show_special_list(self):
        ui.tableWidget_3.setRowCount(0)
        row = 0
        ui.tableWidget_3.setColumnCount(10)
        ui.tableWidget_3.setEditTriggers(QTableWidget.NoEditTriggers)
        ui.tableWidget_3.setHorizontalHeaderLabels(
            ["勾選", "編號", "姓名", "學號/員編", "正臉照", "左臉照", "右臉照", "正臉照(口罩)", "左臉照(口罩)", "右臉照(口罩)"])
        ui.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        cursor.execute(
            "SELECT id, name, number, img_path1, img_path2, img_path3, img_path4, img_path5, img_path6 FROM special_list")
        record = cursor.fetchall()

        for id, name, number, img_path1, img_path2, img_path3, img_path4, img_path5, img_path6 in record:
            ui.tableWidget_3.insertRow(row)

            checkbox = QCheckBox()
            ui.tableWidget_3.setCellWidget(row, 0, checkbox)

            text = "None"

            ui.tableWidget_3.setItem(row, 1, QTableWidgetItem(str(id)))
            ui.tableWidget_3.setItem(row, 2, QTableWidgetItem(name))
            ui.tableWidget_3.setItem(row, 3, QTableWidgetItem(number))
            ui.tableWidget_3.setItem(row, 4, QTableWidgetItem(img_path1))
            if os.path.basename(img_path2) != "1.jpg":
                ui.tableWidget_3.setItem(row, 5, QTableWidgetItem(img_path2))
            else:
                ui.tableWidget_3.setItem(row, 5, QTableWidgetItem(text))
            if os.path.basename(img_path3) != "1.jpg":
                ui.tableWidget_3.setItem(row, 6, QTableWidgetItem(img_path3))
            else:
                ui.tableWidget_3.setItem(row, 6, QTableWidgetItem(text))
            if os.path.basename(img_path4) != "1.jpg":
                ui.tableWidget_3.setItem(row, 7, QTableWidgetItem(img_path4))
            else:
                ui.tableWidget_3.setItem(row, 7, QTableWidgetItem(text))
            if os.path.basename(img_path5) != "1.jpg":
                ui.tableWidget_3.setItem(row, 8, QTableWidgetItem(img_path5))
            else:
                ui.tableWidget_3.setItem(row, 8, QTableWidgetItem(text))
            if os.path.basename(img_path6) != "1.jpg":
                ui.tableWidget_3.setItem(row, 9, QTableWidgetItem(img_path6))
            else:
                ui.tableWidget_3.setItem(row, 9, QTableWidgetItem(text))

            row += 1
            ui.tableWidget_3.setSelectionBehavior(QTableWidget.SelectRows)

    def update_time(self):
        # 獲取當前時間
        str_now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 更新時間
        ui.time_label.setText(str_now)

    def start_video(self):
        if self.video_thread and self.video_thread.is_alive():
            print("Video processing is already running.")
        else:
            self.video_thread = Thread(target=self.vt.run)
            self.video_thread.start()

    def stop_video(self):
        if self.video_thread and self.video_thread.is_alive():
            # 停止線程的邏輯，這需要視情況而定
            cv2.destroyAllWindows()
            self.vt.stop()
            self.video_thread.join()
            print("Video processing is stopped.")
        else:
            print("No video processing is currently running.")

    def search_date(self):  # 時間段查詢
        ui.tableWidget.setRowCount(0)  # 刪除所有表格資料

        datestartChoice = ui.search_start_dateEdit.date().toString("yyyy-MM-dd")
        selected_hour = ui.comboBox_start_hour.currentText()
        selected_min = ui.comboBox_start_min.currentText()
        datestartChoice = datestartChoice + "-" + selected_hour + "-" + selected_min
        dateendChoice = ui.search_end_dateEdit.date().toString("yyyy-MM-dd")
        selected_hour = ui.comboBox_end_hour.currentText()
        selected_min = ui.comboBox_end_min.currentText()
        dateendChoice = dateendChoice + "-" + selected_hour + "-" + selected_min
        cursor.execute("SELECT id, feature, name,number, image_path, in_time FROM history WHERE in_time BETWEEN %s AND %s",
                       (datestartChoice, dateendChoice))
        self.record = cursor.fetchall()
        if self.record:
            self.show_data(self.record)
        else:
            self.show_error_message("Error", "此時段無人出現")

    def search_key(self):  # id&名稱查詢
        ui.tableWidget.setRowCount(0)  # 刪除所有表格資料
        input_data = ui.search_key_lineEdit.text()
        input_data = input_data.strip()
        if input_data is None:
            self.show_error_message("Error", "請輸入查詢關鍵字")
            return
        elif input_data.isdigit() and len(input_data) == 8:
            column_name = "number"
        elif input_data.isnumeric():
            column_name = "id"
        else:
            column_name = "name"
        if ui.search_key_today_rbtn.isChecked():
            query = f"SELECT * FROM today WHERE {column_name} = %s"
        elif ui.search_key_history_rbtn.isChecked():
            query = f"SELECT id, feature, name,number, image_path, in_time FROM history WHERE {column_name} = %s"
        else:
            self.show_error_message("Error", "請選擇查詢方式")
            return

        cursor.execute(query, (input_data,))
        self.record = cursor.fetchall()
        if self.record:
            self.show_data(self.record)
        else:
            self.show_error_message("Error", "查無此人")

        ui.search_key_lineEdit.clear()

    # 顯示圖片視窗
    def msg_img(self, image_path, strr):
        pho = cv2.imread(image_path)
        if pho is not None:
            # 如果圖片讀取成功
            # 調整圖片大小
            resized_image = cv2.resize(pho, (180, 180))

            # 將 BGR 格式轉換為 RGB 格式
            rgb_image = cv2.cvtColor(resized_image, cv2.COLOR_BGR2RGB)

            # 將 numpy 陣列轉換為 QImage
            q_img = QtGui.QImage(
                rgb_image.data, rgb_image.shape[1], rgb_image.shape[0], QtGui.QImage.Format_RGB888)

            # 創建一個信息框
            msg_box = QtWidgets.QMessageBox()
            # 設定視窗標題
            msg_box.setWindowTitle("顯示圖片")
            msg_box.setIcon(QtWidgets.QMessageBox.Information)
            msg_box.setText(strr)
            # msg_box.setDetailedText("這裡是圖片：")詳細資料按鈕

            # 在 detailed text 中插入 QLabel 以顯示圖片
            label = QtWidgets.QLabel()
            pixmap = QtGui.QPixmap.fromImage(q_img)
            label.setPixmap(pixmap)
            label.setAlignment(QtCore.Qt.AlignCenter)

            # 創建“確定”按鈕
            ok_button = msg_box.addButton(QtWidgets.QMessageBox.Ok)
            ok_button.setText("確定")

            # 設置 "確定" 按鈕的位置
            layout = msg_box.layout()
            layout.setContentsMargins(10, 10, 10, 10)  # 設定邊緣空白，以便更好地控制位置
            layout.addWidget(label, 0, 0, 1, 1, QtCore.Qt.AlignCenter)
            layout.addWidget(ok_button, 1, 2, 1, 1,
                             QtCore.Qt.AlignBottom | QtCore.Qt.AlignRight)
            # layout.addWidget(ok_button, 1, 0, 1, 1, QtCore.Qt.AlignBottom | QtCore.Qt.AlignLeft)

            result = msg_box.exec_()

            if result == QtWidgets.QMessageBox.Ok:
                print("User clicked Ok")
                return True
            elif result == QtWidgets.QMessageBox.Cancel:
                print("User clicked Cancel")
            else:
                return False

        else:
            print("圖片讀取失敗")

    def upload_img(self):  # 圖片上傳
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.search_img_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "圖片上傳成功")

    def search_pic(self):  # 圖片查詢
        ui.tableWidget.setRowCount(0)  # 刪除所有表格資料
        if self.pho is None:
            self.show_error_message("Error", "請上傳正臉照")
            ui.search_img_lineEdit.clear()
            return
        box = self.face_detector.detect(self.pho)
        if box[0] is not None and len(box) > 0:
            for face_data in box:
                if len(face_data[0]) >= 4:
                    bbox_xyxy = [int(coord) for coord in face_data[0][:4]]
                    encoding = self.vt.encoding(self.pho, bbox_xyxy)
                    num = self.vt.recognize_face(
                        encoding, special_encoding, special_number)
                    if num in special_number:
                        name = special_name[special_number.index(num)]
                    else:
                        name = None
                    print(name)
                    print(num)
                    break
        else:
            self.show_error_message("Error", "請上傳正臉照")
            self.pho = None
            ui.search_img_lineEdit.clear()
            return
        if num:
            if ui.search_img_history_rbtn.isChecked():
                cursor.execute(
                    "SELECT * FROM today WHERE number = %s ", (num,))

            elif ui.search_img_today_rbtn.isChecked():
                cursor.execute(
                    "SELECT id, feature, name,number, image_path, in_time FROM history WHERE number = %s ", (num,))
            else:
                self.show_error_message("Error", "請選擇查詢方式")
                return
            self.record = cursor.fetchall()

            if self.record:
                print("Following Records:")
                self.show_data(self.record)
                self.msg_img(self.record[0][4], "查詢成功")
            else:
                self.show_error_message("Error", "查無此人")
        else:
            self.show_error_message("Error", "查無此人")

        ui.search_img_lineEdit.clear()

    def specialdata(self):  # 特殊名單
        if ui.search_only_special_ck.isChecked():
            ui.tableWidget.setRowCount(0)
            self.show_data(self.record)
        elif not ui.search_only_special_ck.isChecked():
            ui.tableWidget.setRowCount(0)
            self.show_data(self.record)

    def check(self):
        selected_items = ui.tableWidget.selectedItems()
        selected_rows = set()
        for item in selected_items:
            selected_rows.add(item.row())
        for row in selected_rows:
            image_path = ui.tableWidget.item(row, 4)
            print(image_path.text())
            self.msg_img(image_path.text(), "")

    def check_special(self):
        selected_items = ui.tableWidget_3.selectedItems()  # 獲取所有選中的單元格
        selected_rows = set()
        img_str = ["正臉照", "左側臉照", "右側臉照", "(口罩)正臉照", "(口罩)左側臉照", "(口罩)右側臉照"]
        for item in selected_items:
            selected_rows.add(item.row())  # 獲取選中單元格的行號
        for row in selected_rows:
            name_item = ui.tableWidget_3.item(row, 2)
            name = name_item.text()
            for i in range(4, 10):
                strr = ""
                image_path_item = ui.tableWidget_3.item(row, i)
                if image_path_item.text() == "None":
                    image_path = "./icon/Nobody.jpg"
                    strr = "未上傳"+name+"的" + img_str[i-4]
                else:
                    image_path = image_path_item.text()
                    strr = name + "的" + img_str[i-4]
                print(image_path)
                self.msg_img(image_path, strr)

    def insert_check_specialdata(self):  # 新增勾選特殊名單
        data = []
        name = []
        number = []
        know = True
        count = 0
        for row in range(ui.tableWidget.rowCount()):
            # 獲取該列中的複選框
            checkbox_item = ui.tableWidget.cellWidget(row, 0)
            if checkbox_item and checkbox_item.isChecked():  # 複選框被勾選
                know = False
                count += 1
                ph = ui.tableWidget.item(row, 4).text()
                parent_window = QWidget()
                # 呼叫getText方法顯示輸入對話框
                self.msg_img(ph, "新增此人到特殊名單")
                # 學號
                text, ok = QInputDialog.getText(
                    parent_window, '新增名單', '請輸入學號：')
                if ok:
                    if text == "" or len(text) != 8:
                        self.show_error_message("Error", "請輸入學號")
                        continue
                    elif text in special_number:
                        self.show_error_message("Error", "名單已存在")
                        continue
                    else:
                        number.append(text)
                    # 名字
                    text, ok = QInputDialog.getText(
                        parent_window, '新增名單', '請輸入姓名：')
                    if ok:
                        if text == "":
                            self.show_error_message("Error", "請輸入姓名")
                            continue
                        else:
                            name.append(text)

                    item = ui.tableWidget.item(row, 1)  # 獲取第一列的數據
                    data.append(item.text())

        if know:
            self.show_error_message("Error", "請勾選名單")
            return
        for d, n, num in zip(data, name, number):
            cursor.execute("SELECT * FROM history WHERE id = %s ", (d,))
            record = cursor.fetchall()
            # 新增特殊名單資料夾
            new_path = "./special_list/"+num
            if not os.path.exists(new_path):
                os.mkdir(new_path)
            new_path = os.path.join(new_path, '1.jpg')
            shutil.copy(record[0][4], new_path)
            ui.tableWidget.setRowCount(0)   # 刪除所有表格資料
            self.show_data(
                [(d, record[0][1], n, num, record[0][4], record[0][5])])
            self.msg_img(new_path, "新增成功")

            insert_query = "INSERT INTO special_list (name,number,feature1,img_path1,feature2,img_path2,feature3,img_path3,feature4,img_path4,feature5,img_path5,feature6,img_path6) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            cursor.execute(insert_query, (n, num, record[0][1], new_path, record[0][1], new_path, record[0]
                           [1], new_path, record[0][1], new_path, record[0][1], new_path, record[0][1], new_path))

            insert_query = "UPDATE history SET name = %s, number = %s WHERE id = %s"
            cursor.execute(insert_query, (n, num, d))

            insert_query = "UPDATE today SET name = %s, number=%s WHERE id = %s"
            cursor.execute(insert_query, (n, num, d))
            conn.commit()

            for i in range(6):
                special_encoding.append(pickle.loads(record[0][1]))
                special_name.append(n)
                special_number.append(num)
        if count != len(name):
            self.show_error_message(
                "Error", "有 {} 筆資料建立失敗".format(count-len(name)))

    def show_data(self, record):
        row = 0
        ui.tableWidget.setColumnCount(6)    # 設置欄位數
        ui.tableWidget.setEditTriggers(QTableWidget.NoEditTriggers)
        ui.tableWidget.setHorizontalHeaderLabels(
            ["勾選", "ID", "Name", "number", "Image Path", "In Time"])  # 設置標題

        if self.record is None:
            return

        for id, feature, name, number, image_path, in_time in record:
            if ui.search_only_special_ck.isChecked():  # 選定特殊名單
                if name is None:
                    continue

            ui.tableWidget.insertRow(row)  # 插入一行

            checkbox = QCheckBox()  # 勾選框
            ui.tableWidget.setCellWidget(row, 0, checkbox)  # 插入勾选框

            ui.tableWidget.setItem(row, 1, QTableWidgetItem(str(id)))  # id
            ui.tableWidget.setItem(row, 2, QTableWidgetItem(name))  # name
            ui.tableWidget.setItem(row, 3, QTableWidgetItem(number))  # name
            ui.tableWidget.setItem(
                row, 4, QTableWidgetItem(image_path))  # image_path
            # in_time
            formatted_time = in_time.strftime('%Y-%m-%d %H:%M:%S')
            time_item = QTableWidgetItem(formatted_time)
            time_item.setTextAlignment(Qt.AlignCenter)  # 居中對齊
            ui.tableWidget.setItem(row, 5, time_item)
            row += 1
        ui.tableWidget.setSelectionBehavior(QTableWidget.SelectRows)

     # 統計設定日期選擇(日、週、月、年)

    def handle_date_time_change(self):
        date_start = ui.data_start_dateEdit.date()
        date_end = ui.data_end_dateEdit.date()
        days_difference = date_start.daysTo(date_end)
        print(f"Days Difference: {days_difference} days")
        # 清空下拉式選單
        ui.data_timeunit_cob.clear()

        # 根據下拉式選單的值進行不同的操作
        if days_difference < 7:  # 日
            ui.data_timeunit_cob.addItems(["日"])  # 日
            ui.data_timeunit_cob.setCurrentIndex(0)

        elif days_difference < 31:  # 週
            ui.data_timeunit_cob.addItems(["日", "週"])  # 日、週
            ui.data_timeunit_cob.setCurrentIndex(0)
        elif days_difference < 365:  # 月
            ui.data_timeunit_cob.addItems(["週", "月"])  # 週、月
            ui.data_timeunit_cob.setCurrentIndex(0)
        else:
            ui.data_timeunit_cob.addItems(["月", "年"])  # 月、年
            ui.data_timeunit_cob.setCurrentIndex(0)

    def create_chart_data(self, data_start, data_end):  # 處理資料庫
        if data_start >= data_end:
            self.show_error_message("Error", "起始日期不能大於等於結束日期")
            return 0
        cursor.execute(
            "SELECT * FROM people_count WHERE date BETWEEN %s AND %s", (data_start, data_end))
        self.choose_record = cursor.fetchall()
        print(data_start, data_end)
        daily_data = {}
        for date, count in self.choose_record:
            daily_data[date.strftime("%Y-%m-%d")] = count
        return daily_data

    def chart_data_process(self, data):  # 處理資料(日、週、月、年)
        curr = {}
        data = sorted(data.items(), key=lambda item: QDateTime.fromString(
            item[0], "yyyy-MM-dd"))
        data_iterator = iter(data)
        first_date, num = next(data_iterator)
        start = QDateTime.fromString(first_date, "yyyy-MM-dd")

        if ui.data_timeunit_cob.currentText() == "日":
            curr[start.toString("yyyy-MM-dd")] = num
            for date, count in data_iterator:
                curr[date] = count
            return curr
        elif ui.data_timeunit_cob.currentText() == "週":
            end = start.addDays(6)
            for date, count in data_iterator:
                date = QDateTime.fromString(date, "yyyy-MM-dd")
                num += count
                if date >= end:
                    curr[start.toString("yyyy-MM-dd")] = num
                    start = date.addDays(1)
                    end = start.addDays(6)
                    num = 0
            curr[start.toString("yyyy-MM-dd")] = num
            return curr
        elif ui.data_timeunit_cob.currentText() == "月":
            end = start.addDays(30)

            for date, count in data_iterator:
                date = QDateTime.fromString(date, "yyyy-MM-dd")
                num += count
                if date >= end:
                    curr[start.toString("yyyy-MM-dd")] = num
                    start = date.addDays(1)
                    end = start.addDays(30)
                    num = 0
            curr[start.toString("yyyy-MM-dd")] = num
            return curr
        else:
            end = start.addDays(364)

            for date, count in data_iterator:
                date = QDateTime.fromString(date, "yyyy-MM-dd")

                num += count
                if date >= end:
                    curr[start.toString("yyyy-MM-dd")] = num
                    start = date.addDays(1)
                    end = start.addDays(364)
                    num = 0

            curr[start.toString("yyyy-MM-dd")] = num
            return curr

    def raddio_chart(self):  # 選擇長條/折線
        date_start = ui.data_start_dateEdit.date()
        date_end = ui.data_end_dateEdit.date()
        print(f"Start Date: {date_start.toString('yyyy-MM-dd')}")
        print(f"End Date: {date_end.toString('yyyy-MM-dd')}")
        # 在初始化函數中，設置RadioButton為不獨佔
        ui.barchart_rbtn.setAutoExclusive(False)
        ui.linechart_rbtn.setAutoExclusive(False)

        get = self.create_chart_data(date_start.toString(
            "yyyy-MM-dd"), date_end.toString("yyyy-MM-dd"))  # 處理資料庫
        if get == 0:
            return
        if not get:
            self.show_error_message("Error", "這段時間沒有資料")
            return
        data = self.chart_data_process(get)
        if ui.barchart_rbtn.isChecked():   # 選擇長條圖
            self.create_bar_chart(data)
            self.chart_type = "bar"
            ui.barchart_rbtn.setChecked(False)

        elif ui.linechart_rbtn.isChecked():  # 選擇折線圖
            self.create_line_chart(data)
            self.chart_type = "line"
            ui.linechart_rbtn.setChecked(False)
        else:
            self.show_error_message("Error", "請選擇圖表類型")
            return
        # 恢復RadioButton的獨佔模式
        ui.barchart_rbtn.setAutoExclusive(True)
        ui.linechart_rbtn.setAutoExclusive(True)

    def clean_chart(self):  # 清除圖表
        if self.chart.series():
            series_list = self.chart.series()  # 清空圖表
            for series in series_list:
                self.chart.removeSeries(series)
            for axis in self.chart.axes():
                self.chart.removeAxis(axis)
            # 清空布局
            while self.new_layout.count():
                item = self.new_layout.takeAt(0)  # 取得佈局中的第一個項目
                widget = item.widget()   # 取得項目中的 widget
                if widget is not None:
                    widget.deleteLater()  # 刪除 widget

    def create_line_chart(self, daily_data):
        # 設定圖表的基本屬性
        self.clean_chart()
        self.chart.setTitle("折線圖")
        self.chart.setAnimationOptions(QChart.NoAnimation)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignBottom)

        # 創建折線圖數列
        self.series = QLineSeries()
        self.series.setName("每日人數")
        self.series.setPen(QPen(QColor("blue"), 3, Qt.SolidLine))

        # 創建散點圖數列
        scatter_series = QScatterSeries()
        scatter_series.setName("每日人數")
        scatter_series.setMarkerSize(10)  # 設定標記的大小
        scatter_series.setBrush(QBrush(Qt.red))  # 設定標記的填充顏色

        # 創建X軸和Y軸
        self.xAxis = QDateTimeAxis()
        self.xAxis.setLabelsAngle(45)
        self.xAxis.setFormat("yyyy/MM/dd")
        self.xAxis.setTitleText("日期")
        self.chart.addAxis(self.xAxis, Qt.AlignBottom)

        self.yAxis = QValueAxis()
        self.yAxis.setTitleText("人數")
        self.yAxis.setLabelFormat("%.0f")
        self.chart.addAxis(self.yAxis, Qt.AlignLeft)

        # 調整字體大小
        font = self.xAxis.labelsFont()
        font.setPointSize(14)
        self.xAxis.setLabelsFont(font)
        font1 = self.yAxis.labelsFont()
        font1.setPointSize(18)
        self.yAxis.setLabelsFont(font1)

        # 將每日的人數添加到折線圖數列中
        for category, value in daily_data.items():
            print(category, value)
            date = QDateTime(QDate.fromString(category, "yyyy-MM-dd"))
            point = QPointF(date.toMSecsSinceEpoch(), value)
            self.series.append(point)
            scatter_series.append(point)

        self.chart.addSeries(self.series)
        self.chart.addSeries(scatter_series)

        self.series.attachAxis(self.xAxis)
        self.series.attachAxis(self.yAxis)
        scatter_series.attachAxis(self.xAxis)
        scatter_series.attachAxis(self.yAxis)

        # 設定Y軸和X軸的範圍
        min_date = QDateTime(QDate.fromString(
            min(daily_data.keys()), "yyyy-MM-dd"))
        max_date = QDateTime(QDate.fromString(
            max(daily_data.keys()), "yyyy-MM-dd"))
        self.yAxis.setRange(int(min(daily_data.values())),
                            int(max(daily_data.values())))
        self.xAxis.setRange(min_date, max_date)
        self.xAxis.setTickCount(len(daily_data))  # 設定刻度的數量

        # 設定點標籤
        self.series.setPointLabelsFormat("@yPoint 人")
        self.series.setPointLabelsVisible(True)
        self.series.setPointLabelsClipping(False)
        self.chart.legend().markers(self.series)[0].setVisible(False)
        # 創建圖表視圖並設定屬性
        self.chart_view = QChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.Antialiasing)
        self.chart_view.setSizePolicy(
            QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 將圖表視圖添加到佈局
        self.new_layout.addWidget(self.chart_view)
        self.setLayout(self.new_layout)

    def create_bar_chart(self, daily_data):
        # 設置圖表屬性
        self.clean_chart()
        self.chart.setTitle("長條圖")
        self.chart.setAnimationOptions(QChart.NoAnimation)
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignBottom)

        # 創建柱狀圖數列（QBarSeries）
        series = QBarSeries()
        # 將每日的人數添加到柱狀圖數列中
        for day, count in daily_data.items():
            bar_set = QBarSet(str(day))
            bar_set.append(count)
            series.append(bar_set)

        # 創建柱狀圖分類軸（QBarCategoryAxis）和值軸（QValueAxis）
        x_axis = QBarCategoryAxis()

        # 調整字體大小
        font = x_axis.labelsFont()
        font.setPointSize(14)
        x_axis.setLabelsFont(font)

        self.chart.setAxisX(x_axis, series)

        counts = [count for day, count in daily_data.items()]
        equal = all(count == counts[0] for count in counts)

        y_axis = QValueAxis()
        # 調整字體大小
        font1 = y_axis.labelsFont()
        font1.setPointSize(18)
        y_axis.setLabelsFont(font1)
        y_axis.setLabelFormat("%.0f")
        if equal:
            y_axis.setRange(counts[0], counts[0] + 3)

            self.chart.setAxisY(y_axis, series)
        else:
            y_axis.setRange(min(daily_data.values()), max(daily_data.values()))
            self.chart.setAxisY(y_axis, series)

        # 將柱狀圖數列添加到圖表
        self.chart.addSeries(series)

        # 創建圖表視圖
        self.chart_view = QChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.Antialiasing)

        # 添加到布局
        self.new_layout.addWidget(self.chart_view)

    def data_export(self):
        # 新增時間單位下拉式選單
        export_units = ["CSV", "Excel", "Image", "PDF"]
        ui.export_cob.addItems(export_units)

    def data_export_btn(self):
        selected_item = ui.export_cob.currentText()
        if selected_item == "CSV":
            self.export_csv()
        elif selected_item == "PDF":
            self.export_PDF()
        elif selected_item == "Excel":
            self.export_Excel()
        elif selected_item == "Image":
            self.export_Image()

    def export_csv(self):
        file_path = 'export_sql_data.csv'
        with open(file_path, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            header = ["date", "count"]
            csv_writer.writerow(header)  # 設表頭
            csv_writer.writerows(self.choose_record)

    def export_Excel(self):
        file_path = 'export_sql_data.xlsx'
        workbook = openpyxl.Workbook()  # 創新excel
        worksheet = workbook.active
        header = ["date", "count"]
        worksheet.append(header)  # 設表頭
        for record in self.choose_record:  # 寫檔
            worksheet.append(record)
        workbook.save(file_path)  # 儲存

    def export_Image(self):
        if self.chart_type == "bar":
            file_path = 'bar_chart.jpg'
        else:
            file_path = 'line_chart.jpg'

        image = QImage(self.chart_view.size(), QImage.Format_ARGB32)
        image.fill(Qt.transparent)
        painter = QPainter(image)
        self.chart_view.render(painter)
        painter.end()
        image.save(file_path)

    def export_PDF(self):
        merger = PdfMerger()

        for page_number in range(2):
            filename = f'page_{page_number + 1}.pdf'
            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []

            if page_number == 0:
                image_filename = f'chart_image_{page_number}.png'
                image = QImage(self.chart_view.size(), QImage.Format_ARGB32)
                image.fill(Qt.transparent)
                painter = QPainter(image)
                self.chart_view.render(painter)
                image.save(image_filename)
                painter.end()

                elements.append(
                    Image(image_filename, width=6*inch, height=4*inch))

            else:
                data = self.choose_record
                data.insert(0, ['date', 'count'])
                table = Table(data)

                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)

            doc.build(elements)
            merger.append(filename)

        merger.write('export_data.pdf')
        merger.close()
        os.remove('page_1.pdf')
        os.remove('page_2.pdf')
        os.remove('chart_image_0.png')

    def show_error_message(self, strr, strr1):  # 警示框
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(strr1)
        msg.setWindowTitle(strr)
        msg.exec_()

    def special_img(self):  # 特定正臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_img_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "正臉照上傳成功")
        # pho=cv2.imread(image_path)

    def special_img_left(self):  # 特定左臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_imgl_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "左側臉照上傳成功")

    def special_img_right(self):  # 特定右臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_imgr_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "右側臉照上傳成功")

    def special_img_mask(self):  # 特定口罩正臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_imgm_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "正臉照(口罩)上傳成功")

    def special_img_mask_left(self):  # 特定口罩左臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_imgml_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "左側臉照(口罩)上傳成功")

    def special_img_mask_right(self):  # 特定口右罩臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        ui.special_imgmr_lineEdit.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "右側臉照(口罩)上傳成功")

    def insert_notify_list(self):
        input_name = ui.notify_name_lineEdit.text()
        input_email = ui.notify_email_lineEdit.text()
        if input_name == "":
            strr = "請輸入姓名!"
            self.show_error_message("未輸入姓名", strr)
        if input_email == "":
            strr = "請輸入電子郵件!"
            self.show_error_message("未輸入電子郵件", strr)

        if input_name and input_email:
            insert_query = "INSERT INTO notify (name, email) VALUES (%s, %s)"
            cursor.execute(insert_query, (input_name, input_email))
            conn.commit()
            notify_email.append(input_email)
            notify_name.append(input_name)

        ui.notify_name_lineEdit.clear()
        ui.notify_email_lineEdit.clear()

        self.show_notify_list()

    def modify_notify_list(self):
        for row in range(ui.tableWidget_2.rowCount()):
            checkbox_item = ui.tableWidget_2.cellWidget(row, 0)
            id = ui.tableWidget_2.item(row, 1).text()
            name = ui.tableWidget_2.item(row, 2).text()
            email = ui.tableWidget_2.item(row, 3).text()
            if checkbox_item and checkbox_item.isChecked():
                parent_window = QWidget()
                name_change, name_ok = QInputDialog.getText(
                    parent_window, "Name", "變更"+name+"的姓名")
                email_change, email_ok = QInputDialog.getText(
                    parent_window, "Email", "變更Email")

                if name_ok and email_ok:
                    if name_change == "":
                        name_change = name
                    else:
                        for i in range(len(notify_name)):
                            if notify_name[i] == name:
                                notify_name[i] = name_change

                    if email_change == "":
                        email_change = email
                    else:
                        for i in range(len(notify_email)):
                            if notify_email[i] == email:
                                notify_email[i] = email_change

                    update_query = "UPDATE notify SET name = %s, email = %s WHERE id = %s"
                    cursor.execute(
                        update_query, (name_change, email_change, id))
                    conn.commit()

        self.show_notify_list()

    def delete_notify_list(self):
        for row in range(ui.tableWidget_2.rowCount()):
            checkbox_item = ui.tableWidget_2.cellWidget(row, 0)
            id = ui.tableWidget_2.item(row, 1).text()
            name = ui.tableWidget_2.item(row, 2).text()
            email = ui.tableWidget_2.item(row, 3).text()
            if checkbox_item and checkbox_item.isChecked():
                del_query = "DELETE FROM notify WHERE id = %s"
                cursor.execute(del_query, (id,))
                conn.commit()

                print(name)
                notify_name.remove(name)
                notify_email.remove(email)

        self.show_notify_list()

    def extract_special_feature(self, name, number, img_pth1, img_pth2, img_pth3, img_pth4, img_pth5, img_pth6):
        folder_path = "./special_list/"+number
        try:
            os.mkdir(folder_path)
        except FileExistsError:
            print(f"The folder {folder_path} already exists.")

        shutil.copy(img_pth1, os.path.join(folder_path, "1.jpg"))

        if img_pth2:
            shutil.copy(img_pth2, os.path.join(folder_path, "2.jpg"))
        if img_pth3:
            shutil.copy(img_pth3, os.path.join(folder_path, "3.jpg"))
        if img_pth4:
            shutil.copy(img_pth4, os.path.join(folder_path, "4.jpg"))
        if img_pth5:
            shutil.copy(img_pth5, os.path.join(folder_path, "5.jpg"))
        if img_pth6:
            shutil.copy(img_pth6, os.path.join(folder_path, "6.jpg"))

        person_data = {
            'name': name,
            'number': number,
            'feature1': None,
            'img_path1': None,
            'feature2': None,
            'img_path2': None,
            'feature3': None,
            'img_path3': None,
            'feature4': None,
            'img_path4': None,
            'feature5': None,
            'img_path5': None,
            'feature6': None,
            'img_path6': None
        }

        '''
        if number:
            person_data["number"] = number
        '''

        for i in range(1, 7):
            img_path = os.path.join(folder_path, f"{i}.jpg")
            if not os.path.exists(img_path):
                # 如果文件不存在，使用第一张照片填充
                img_path = os.path.join(folder_path, "1.jpg")

            image = cv2.imread(img_path)

            faces = self.vt.face_detector.detect(image)

            if len(faces) == 1:
                face = faces[0]  # 假設每張照片只有一個人臉
                x, y, w, h = face['box']
                face_image = image[y:y+h, x:x+w]
            else:
                face_image = image

        # 將人臉圖像轉換為FaceNet所需的格式
            face_image = cv2.cvtColor(face_image, cv2.COLOR_BGR2RGB)
            face_image = cv2.resize(face_image, (160, 160))
            face_image = (face_image - 127.5) / 128.0  # 歸一化

            # 提取人臉特徵
            face_tensor = np.expand_dims(
                face_image.transpose(2, 0, 1), axis=0)
            face_tensor = torch.from_numpy(face_tensor).float()
            features = self.vt.facenet(face_tensor)

            # 更新字典中的特徵和圖像路徑
            person_data[f'feature{i}'] = pickle.dumps(
                features.detach().numpy())
            person_data[f'img_path{i}'] = img_path

            special_encoding.append(pickle.dumps(
                features.detach().numpy()))
            special_name.append(name)

        insert_query = f"INSERT INTO special_list ("
        for key in person_data.keys():
            insert_query += key + ","
        insert_query = insert_query[:-1] + ") VALUES ("
        for key in person_data.keys():
            insert_query += "%s,"
        insert_query = insert_query[:-1] + ")"
        values = tuple(person_data.values())
        cursor.execute(insert_query, values)
        conn.commit()

    def insert_special_xlsx(self, name, number):

        excel = openpyxl.load_workbook('./special_list/special_list.xlsx')
        s1 = excel['工作表1']

        data = []
        folder_path = './special_list/' + number

        data.append(number)
        data.append(name)

        for i in range(1, 7):
            img_path = os.path.join(folder_path, f"{i}.jpg")
            if not os.path.exists(img_path):
                # 如果文件不存在，使用第一张照片填充
                img_path = os.path.join(folder_path, "1.jpg")

            data.append(img_path)

        s1.append(data)
        excel.save('./special_list/special_list.xlsx')

    def modify_special_xlsx(self, number, data):
        excel = openpyxl.load_workbook('./special_list/special_list.xlsx')
        s1 = excel['工作表1']  # 获取名为 "工作表1" 的工作表

        index = 0

        for row in s1.iter_rows(values_only=True):
            if row[0] == number:
                break
            else:
                index += 1

        for i, value in enumerate(data):
            value = str(value)
            s1.cell(row=index + 1, column=i + 1, value=value)

        excel.save('./special_list/special_list.xlsx')

    def delete_special_xlsx(self, name):

        # Load the Excel file
        excel = openpyxl.load_workbook('./special_list/special_list.xlsx')

        # Get the worksheet
        worksheet = excel['工作表1']

        # List to store the rows to be deleted
        rows_to_delete = []

        # Loop through the rows
        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            # If the name is found, add the row index to the delete list
            if name in row:
                rows_to_delete.append(i)

        # Loop through the rows to be deleted and delete them
        for i in rows_to_delete:
            worksheet.delete_rows(i + 1)  # Add 1, as row indices start from 1

        # Save the Excel file
        excel.save('./special_list/special_list.xlsx')

    def insert_special_list(self):

        name = ui.special_name_lineEdit.text()
        number = ui.special_num_lineEdit.text()
        img_pth1 = ui.special_img_lineEdit.text()
        img_pth2 = ui.special_imgl_lineEdit.text()
        img_pth3 = ui.special_imgr_lineEdit.text()
        img_pth4 = ui.special_imgm_lineEdit.text()
        img_pth5 = ui.special_imgml_lineEdit.text()
        img_pth6 = ui.special_imgmr_lineEdit.text()

        if name == "":
            strr = "請輸入姓名!"
            self.show_error_message("未輸入姓名", strr)

        if number == "":
            strr = "請輸入學號/員編!"
            self.show_error_message("未輸入學號/員編!", strr)

        if img_pth1 == "":
            strr = "請上傳正臉照!"
            self.show_error_message("未輸入正臉照", strr)

        if name and number and img_pth1:
            self.extract_special_feature(
                name, number, img_pth1, img_pth2, img_pth3, img_pth4, img_pth5, img_pth6)
            self.insert_special_xlsx(
                name, number)

        # print(special_name)
        # print(special_encoding)

        ui.special_name_lineEdit.clear()
        ui.special_num_lineEdit.clear()
        ui.special_img_lineEdit.clear()
        ui.special_imgl_lineEdit.clear()
        ui.special_imgr_lineEdit.clear()
        ui.special_imgm_lineEdit.clear()
        ui.special_imgml_lineEdit.clear()
        ui.special_imgmr_lineEdit.clear()

        self.show_special_list()

    def delete_special_list(self):
        for row in range(ui.tableWidget_3.rowCount()):
            checkbox_item = ui.tableWidget_3.cellWidget(row, 0)
            number = ui.tableWidget_3.item(row, 3).text()
            if checkbox_item and checkbox_item.isChecked():
                folder_path = "./special_list/"+number
                try:
                    shutil.rmtree(folder_path)
                   # print(f"資料夾 {folder_path} 已刪除，包括所有子目錄和檔案")
                    print('delete finish')
                except OSError as e:
                    # print(f"刪除資料夾時出現錯誤: {e}")
                    print(f"error: {e}")

                try:
                    cursor.execute(
                        "SELECT * FROM special_list WHERE number = %s", (number,))
                    result = cursor.fetchone()
                    conn.commit()
                except mysql.connector.Error as mysql_error:
                    print(f"MySQL error: {mysql_error}")

                if result:
                    id, name, number, feature1, img_pth1, feature2, img_pth2, feature3, img_pth3, feature4, img_pth4, feature5, img_pth5, feature6, img_pth6 = result
                else:
                    print("未找到匹配的数据")
                elements_to_remove = [feature1, feature2,
                                      feature3, feature4, feature5, feature6]

                for element in elements_to_remove:
                    serialized_element = element  # 將要刪除的元素序列化
                    while serialized_element in special_encoding:
                        special_encoding.remove(serialized_element)

                for i in range(1, 7):
                    special_name.remove(name)

                # print(special_name)

                del_query = "DELETE FROM special_list WHERE name = %s"
                cursor.execute(del_query, (name,))
                conn.commit()

                self.delete_special_xlsx(name)

        self.show_special_list()

    def showDialog(self):
        for row in range(ui.tableWidget_3.rowCount()):
            checkbox_item = ui.tableWidget_3.cellWidget(row, 0)
            name = ui.tableWidget_3.item(row, 2).text()
            if checkbox_item and checkbox_item.isChecked():
                self.dialog = QDialog()
                self.di = Ui_Dialog()
                self.di.setupUi(self.dialog)
                cursor.execute(
                    "SELECT * FROM special_list Where name = %s", (name,))
                result = cursor.fetchone()
                conn.commit()

                if result:
                    id, name, number, img_pth1, feature1, img_pth2, feature2, img_pth3, feature3, img_pth4, feature4, img_pth5, feature5, img_pth6, feature6 = result
                else:
                    print("未找到匹配的数据")

                self.di.buttonBox.accepted.connect(lambda: self.modify_special_list(
                    id, name, number, img_pth1, feature1, img_pth2, feature2, img_pth3, feature3, img_pth4, feature4, img_pth5, feature5, img_pth6, feature6))
                self.di.buttonBox.rejected.connect(self.dialog.reject)
                self.di.pushButton.clicked.connect(self.modify_special_img)
                self.di.pushButton_2.clicked.connect(
                    self.modify_special_img_left)
                self.di.pushButton_3.clicked.connect(
                    self.modify_special_img_right)
                self.di.pushButton_4.clicked.connect(
                    self.modify_special_img_mask)
                self.di.pushButton_5.clicked.connect(
                    self.modify_special_img_mask_left)
                self.di.pushButton_6.clicked.connect(
                    self.modify_special_img_mask_right)
                self.dialog.exec_()

    def modify_special_img(self):  # 特定正臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_3.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "正臉照上傳成功")
        # pho=cv2.imread(image_path)
        # print("modify_special_img 函数被调用")

    def modify_special_img_left(self):  # 特定左臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_4.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "左側照上傳成功")

    def modify_special_img_right(self):  # 特定右臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_5.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "右側照上傳成功")

    def modify_special_img_mask(self):  # 特定口罩正臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_6.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "(口罩)正臉照上傳成功")

    def modify_special_img_mask_left(self):  # 特定口罩左臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_7.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "(口罩)左側照上傳成功")

    def modify_special_img_mask_right(self):  # 特定口右罩臉
        image_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "上傳", "./")
        self.di.lineEdit_8.setText(image_path)
        self.pho = cv2.imread(image_path)
        self.msg_img(image_path, "(口罩)右側照上傳成功")

    def detect_encoding(self, img_pth):
        img = cv2.imread(img_pth)
        box = self.face_detector.detect(img)
        if box[0] is not None and len(box) > 0:
            for face_data in box:
                if len(face_data[0]) >= 4:
                    bbox_xyxy = [int(coord) for coord in face_data[0][:4]]
                    encoding = self.vt.encoding(img, bbox_xyxy)
                    return encoding

    def modify_special_list(self, id, name, number, img_pth1, feature1, img_pth2, feature2, img_pth3, feature3, img_pth4, feature4, img_pth5, feature5, img_pth6, feature6):
        name_m = self.di.lineEdit.text()
        number_m = self.di.lineEdit_2.text()
        img_pth1_m = self.di.lineEdit_3.text()
        img_pth2_m = self.di.lineEdit_4.text()
        img_pth3_m = self.di.lineEdit_5.text()
        img_pth4_m = self.di.lineEdit_6.text()
        img_pth5_m = self.di.lineEdit_7.text()
        img_pth6_m = self.di.lineEdit_8.text()

        folder_path = "./special_list/"

        index = 0

        data = [number, name, img_pth1, img_pth2,
                img_pth3, img_pth4, img_pth5, img_pth6]

        for i in range(len(special_name)):
            if special_name[i] == name:
                index = i
                break

        if number_m == "":
            number_m = number
            folder_path = folder_path+number_m
        else:
            update_query = "UPDATE special_list SET number = %s WHERE id = %s"
            cursor.execute(update_query, (number_m, id))
            os.rename(folder_path+number, folder_path+number_m)
            folder_path = folder_path+number_m
            data[0] = number_m

            for i in range(1, 7):
                img_path = os.path.join(folder_path, f"{i}.jpg")
                if not os.path.exists(img_path):
                    img_path = os.path.join(folder_path, "1.jpg")

                data[i+1] = img_path

                update_query = f"UPDATE special_list SET img_path{i} = %s WHERE id = %s"
                cursor.execute(update_query, (img_path, id))

        if name_m == "":
            name_m = name
        else:
            update_query = "UPDATE special_list SET name = %s WHERE id = %s"
            cursor.execute(update_query, (name_m, id))
            data[1] = name_m
            for i in range(0, 6):
                special_name[index+i] = name_m

        if img_pth1_m == "":
            img_pth1_m = data[2]
            feature1_m = feature1
        else:
            shutil.copy(img_pth1_m, os.path.join(folder_path, "1.jpg"))
            img_pth1_m = os.path.join(folder_path, "1.jpg")
            data[2] = img_pth1_m
            feature1_m = self.detect_encoding(img_pth1_m)
            feature1_m = pickle.dumps(feature1_m.detach().numpy())
            special_encoding[index] = feature1_m
            update_query = "UPDATE special_list SET img_path1 = %s, feature1 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth1_m, feature1_m, id))

        if img_pth2_m == "":
            img_pth2_m = data[3]
            feature2_m = feature2
        else:
            shutil.copy(img_pth2_m, os.path.join(folder_path, "2.jpg"))
            img_pth2_m = os.path.join(folder_path, "2.jpg")
            data[3] = img_pth2_m
            feature2_m = self.detect_encoding(img_pth2_m)
            feature2_m = pickle.dumps(feature2_m.detach().numpy())
            special_encoding[index+1] = feature2_m
            update_query = "UPDATE special_list SET img_path2 = %s, feature2 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth2_m, feature2_m, id))

        if img_pth3_m == "":
            img_pth3_m = data[4]
            feature3_m = feature3
        else:
            shutil.copy(img_pth3_m, os.path.join(folder_path, "3.jpg"))
            img_pth3_m = os.path.join(folder_path, "3.jpg")
            data[4] = img_pth3_m
            feature3_m = self.detect_encoding(img_pth3_m)
            feature3_m = pickle.dumps(feature3_m.detach().numpy())
            special_encoding[index+2] = feature3_m
            update_query = "UPDATE special_list SET img_path3 = %s, feature3 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth3_m, feature3_m, id))

        if img_pth4_m == "":
            img_pth4_m = data[5]
            feature4_m = feature4
        else:
            shutil.copy(img_pth4_m, os.path.join(folder_path, "4.jpg"))
            img_pth4_m = os.path.join(folder_path, "4.jpg")
            data[5] = img_pth4_m
            feature4_m = self.detect_encoding(img_pth4_m)
            feature4_m = pickle.dumps(feature4_m.detach().numpy())
            special_encoding[index+3] = feature4_m
            update_query = "UPDATE special_list SET img_path4 = %s, feature4 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth4_m, feature4_m, id))

        if img_pth5_m == "":
            img_pth5_m = data[6]
            feature5_m = feature5
        else:
            shutil.copy(img_pth5_m, os.path.join(folder_path, "5.jpg"))
            img_pth5_m = os.path.join(folder_path, "5.jpg")
            data[6] = img_pth5_m
            feature5_m = self.detect_encoding(img_pth5_m)
            feature5_m = pickle.dumps(feature5_m.detach().numpy())
            special_encoding[index+4] = feature5_m
            update_query = "UPDATE special_list SET img_path5 = %s, feature5 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth5_m, feature5_m, id))

        if img_pth6_m == "":
            img_pth6_m = data[7]
            feature6_m = feature6
        else:
            shutil.copy(img_pth6_m, os.path.join(folder_path, "6.jpg"))
            img_pth6_m = os.path.join(folder_path, "6.jpg")
            data[7] = img_pth6_m
            feature6_m = self.detect_encoding(img_pth6_m)
            feature6_m = pickle.dumps(feature6_m.detach().numpy())
            special_encoding[index+5] = feature6_m
            update_query = "UPDATE special_list SET img_path6 = %s, feature6 = %s WHERE id = %s"
            cursor.execute(update_query, (img_pth6_m, feature6_m, id))

        conn.commit()

        '''
        self.di.lineEdit.clear()
        self.di.lineEdit_2.clear()
        self.di.lineEdit_3.clear()
        self.di.lineEdit_4.clear()
        self.di.lineEdit_5.clear()
        self.di.lineEdit_6.clear()
        self.di.lineEdit_7.clear()
        self.di.lineEdit_8.clear()
        '''

        print(data)

        self.modify_special_xlsx(number, data)

        self.show_special_list()

        self.dialog.accept()


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow_controller()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
